import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from django.db.models import Count, Q, Avg
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from accounts.permissions import IsAdminOrTeacher, IsAdminUser
from students.models import Student, Grade, Section, TeacherAssignment
from attendance.models import AttendanceRecord

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request):
        user = request.user
        today = timezone.localdate()
        
        # 1. Base Querysets based on roles
        if user.role == 'ADMIN':
            sections = Section.objects.all()
            students = Student.objects.all()
            attendance = AttendanceRecord.objects.all()
        else:
            # Teacher
            assigned_section_ids = TeacherAssignment.objects.filter(teacher=user).values_list('section_id', flat=True)
            sections = Section.objects.filter(id__in=assigned_section_ids)
            students = Student.objects.filter(section_id__in=assigned_section_ids)
            attendance = AttendanceRecord.objects.filter(student__section_id__in=assigned_section_ids)

        total_students_count = students.count()
        if total_students_count == 0:
            return Response({
                "total_students": 0,
                "today_present": 0,
                "today_absent": 0,
                "today_unmarked": 0,
                "today_rate": 0,
                "grade_stats": [],
                "section_stats": [],
                "trend_stats": [],
                "most_absent": []
            })

        # 2. Today's Attendance Overview
        today_records = attendance.filter(date=today)
        today_present = today_records.filter(status='PRESENT').count()
        today_absent = today_records.filter(status='ABSENT').count()
        today_unmarked = max(0, total_students_count - (today_present + today_absent))
        
        today_rate = (today_present / (today_present + today_absent) * 100) if (today_present + today_absent) > 0 else 0

        # 3. Stats by Grade
        grade_stats = []
        grades = Grade.objects.all()
        for grade in grades:
            # Filter students by grade (and assigned sections if teacher)
            grade_students = students.filter(section__grade=grade)
            g_total = grade_students.count()
            if g_total > 0:
                g_present = today_records.filter(student__section__grade=grade, status='PRESENT').count()
                g_absent = today_records.filter(student__section__grade=grade, status='ABSENT').count()
                g_rate = (g_present / (g_present + g_absent) * 100) if (g_present + g_absent) > 0 else 0
                grade_stats.append({
                    "grade_id": grade.id,
                    "grade_name": grade.name,
                    "total_students": g_total,
                    "today_present": g_present,
                    "today_absent": g_absent,
                    "attendance_rate": round(g_rate, 1)
                })

        # 4. Stats by Section
        section_stats = []
        for section in sections:
            sec_students = students.filter(section=section)
            s_total = sec_students.count()
            if s_total > 0:
                s_present = today_records.filter(student__section=section, status='PRESENT').count()
                s_absent = today_records.filter(student__section=section, status='ABSENT').count()
                s_rate = (s_present / (s_present + s_absent) * 100) if (s_present + s_absent) > 0 else 0
                section_stats.append({
                    "section_id": section.id,
                    "section_name": f"{section.grade.name} - {section.name}",
                    "total_students": s_total,
                    "today_present": s_present,
                    "today_absent": s_absent,
                    "attendance_rate": round(s_rate, 1)
                })

        # 5. Historical Trend (Last 7 Days)
        trend_stats = []
        last_7_days = [today - timezone.timedelta(days=i) for i in range(6, -1, -1)]
        for d in last_7_days:
            d_records = attendance.filter(date=d)
            d_present = d_records.filter(status='PRESENT').count()
            d_absent = d_records.filter(status='ABSENT').count()
            d_rate = (d_present / (d_present + d_absent) * 100) if (d_present + d_absent) > 0 else 0
            trend_stats.append({
                "date": d.strftime("%b %d"),
                "present": d_present,
                "absent": d_absent,
                "attendance_rate": round(d_rate, 1)
            })

        # 6. Ranked Most Absent Students (Limit 10)
        # Filter: Students with highest count of ABSENT records
        most_absent_query = students.annotate(
            absent_count=Count('attendance_records', filter=Q(attendance_records__status='ABSENT'))
        ).order_by('-absent_count')[:10]
        
        most_absent = [
            {
                "student_id": s.student_id,
                "full_name": s.full_name,
                "grade": s.section.grade.name,
                "section": s.section.name,
                "absent_count": s.absent_count
            } for s in most_absent_query if s.absent_count > 0
        ]

        # 7. Absent list for today
        absent_today_query = today_records.filter(status='ABSENT').select_related('student', 'student__section', 'student__section__grade')
        absent_today = [
            {
                "student_id": rec.student.student_id,
                "full_name": rec.student.full_name,
                "grade": rec.student.section.grade.name,
                "section": rec.student.section.name,
            } for rec in absent_today_query
        ]

        return Response({
            "total_students": total_students_count,
            "today_present": today_present,
            "today_absent": today_absent,
            "today_unmarked": today_unmarked,
            "today_rate": round(today_rate, 1),
            "grade_stats": grade_stats,
            "section_stats": section_stats,
            "trend_stats": trend_stats,
            "most_absent": most_absent,
            "absent_today": absent_today
        })

class DownloadExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request):
        user = request.user
        
        # Get query parameters
        grade_id = request.GET.get('grade')
        section_id = request.GET.get('section')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        status_param = request.GET.get('status')
        
        # Base Filter
        if user.role == 'ADMIN':
            records = AttendanceRecord.objects.all().select_related('student', 'student__section', 'student__section__grade')
        else:
            assigned_sections = TeacherAssignment.objects.filter(teacher=user).values_list('section_id', flat=True)
            records = AttendanceRecord.objects.filter(student__section_id__in=assigned_sections).select_related('student', 'student__section', 'student__section__grade')

        # Apply Filters
        if grade_id:
            records = records.filter(student__section__grade_id=grade_id)
        if section_id:
            records = records.filter(student__section_id=section_id)
        if start_date:
            records = records.filter(date__gte=start_date)
        if end_date:
            records = records.filter(date__lte=end_date)
        if status_param:
            records = records.filter(status=status_param.upper())

        records = records.order_by('-date', 'student__full_name')

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="1B365D")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # 1. School Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "BURKA HARBU SECONDARY SCHOOL"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 30

        # 2. Report Subtitle
        ws.merge_cells('A2:G2')
        subtitle = "Digital Attendance Management System Report"
        if start_date or end_date:
            subtitle += f" ({start_date or 'Start'} to {end_date or 'End'})"
        ws['A2'] = subtitle
        ws['A2'].font = Font(name="Arial", size=11, italic=True)
        ws['A2'].alignment = align_center
        ws.row_dimensions[2].height = 20

        ws.append([]) # Empty spacing row

        # 3. Table Headers
        headers = ["Student ID", "Full Name", "Grade", "Section", "Date", "Status", "Scanned Timestamp"]
        ws.append(headers)
        header_row_index = 4
        ws.row_dimensions[header_row_index].height = 25
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_index, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        # 4. Populate Data
        row_idx = 5
        for rec in records:
            timestamp_str = timezone.localtime(rec.timestamp).strftime('%Y-%m-%d %H:%M:%S') if rec.timestamp else "N/A"
            row_data = [
                rec.student.student_id,
                rec.student.full_name,
                rec.student.section.grade.name,
                rec.student.section.name,
                rec.date.strftime('%Y-%m-%d'),
                rec.get_status_display(),
                timestamp_str
            ]
            ws.append(row_data)
            
            # Formatting data row
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 0)
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Zebra striping
                if is_zebra:
                    cell.fill = zebra_fill
                    
                # Alignments
                if col_idx in [1, 3, 4, 5, 6, 7]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
                # Highlights for Status
                if col_idx == 6: # Status column
                    if rec.status == 'PRESENT':
                        cell.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
                    else:
                        cell.font = Font(name="Arial", size=10, bold=True, color="C62828")
            row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row > 2 and cell.value:  # Ignore merged header rows for length calculation
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Build Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_name = f"Burka_Harbu_Attendance_{timezone.localdate().isoformat()}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={file_name}"
        
        wb.save(response)
        return response

class DownloadPDFView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request):
        user = request.user
        
        # Get query parameters
        grade_id = request.GET.get('grade')
        section_id = request.GET.get('section')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        status_param = request.GET.get('status')
        
        # Base Filter
        if user.role == 'ADMIN':
            records = AttendanceRecord.objects.all().select_related('student', 'student__section', 'student__section__grade')
        else:
            assigned_sections = TeacherAssignment.objects.filter(teacher=user).values_list('section_id', flat=True)
            records = AttendanceRecord.objects.filter(student__section_id__in=assigned_sections).select_related('student', 'student__section', 'student__section__grade')

        # Apply Filters
        if grade_id:
            records = records.filter(student__section__grade_id=grade_id)
        if section_id:
            records = records.filter(student__section_id=section_id)
        if start_date:
            records = records.filter(date__gte=start_date)
        if end_date:
            records = records.filter(date__lte=end_date)
        if status_param:
            records = records.filter(status=status_param.upper())

        records = records.order_by('-date', 'student__full_name')[:200]  # Cap PDF at 200 items for printing speed/limit

        response = HttpResponse(content_type='application/pdf')
        file_name = f"Burka_Harbu_Attendance_{timezone.localdate().isoformat()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Build PDF Document
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'SchoolTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#1B365D'),
            alignment=1  # Centered
        )
        
        subtitle_style = ParagraphStyle(
            'SchoolSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=11,
            leading=14,
            textColor=colors.HexColor('#D4AF37'), # Gold
            alignment=1  # Centered
        )
        
        summary_style = ParagraphStyle(
            'SummaryText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#333333'),
        )
        
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=1
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#333333'),
            alignment=0
        )
        
        table_cell_center_style = ParagraphStyle(
            'TableCellCenter',
            parent=table_cell_style,
            alignment=1
        )

        elements = []

        # 1. Header Block
        elements.append(Paragraph("BURKA HARBU SECONDARY SCHOOL", title_style))
        elements.append(Spacer(1, 4))
        elements.append(Paragraph("DIGITAL ATTENDANCE MANAGEMENT SYSTEM", subtitle_style))
        elements.append(Spacer(1, 15))

        # 2. Report Details / Statistics Summary
        total_recs = records.count()
        pres_count = records.filter(status='PRESENT').count()
        abs_count = records.filter(status='ABSENT').count()
        rate = (pres_count / total_recs * 100) if total_recs > 0 else 0

        details_html = f"""
        <b>Report Details:</b><br/>
        Date Range: {start_date or 'All Time'} to {end_date or 'Today'}<br/>
        Filter: Grade: {Grade.objects.get(id=grade_id).name if grade_id else 'All'}, Section: {Section.objects.get(id=section_id).name if section_id else 'All'}<br/>
        Generated On: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        """
        
        stats_html = f"""
        <b>Summary Statistics:</b><br/>
        Total Records Shown: {total_recs}<br/>
        Present: {pres_count} &nbsp;&nbsp;&nbsp; Absent: {abs_count}<br/>
        Attendance Rate: {rate:.1f}%<br/>
        """

        summary_data = [
            [Paragraph(details_html, summary_style), Paragraph(stats_html, summary_style)]
        ]
        
        summary_table = Table(summary_data, colWidths=[270, 270])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8F9FA')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('PADDING', (0,0), (-1,-1), 10),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E0E0E0')),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # 3. Attendance Table
        table_data = [
            [
                Paragraph("Student ID", table_header_style), 
                Paragraph("Full Name", table_header_style), 
                Paragraph("Grade", table_header_style), 
                Paragraph("Section", table_header_style), 
                Paragraph("Date", table_header_style), 
                Paragraph("Status", table_header_style), 
                Paragraph("Timestamp", table_header_style)
            ]
        ]

        for idx, rec in enumerate(records):
            timestamp_str = timezone.localtime(rec.timestamp).strftime('%H:%M:%S') if rec.timestamp else "—"
            
            # Status styling
            status_color = '#2E7D32' if rec.status == 'PRESENT' else '#C62828'
            status_text = f"<b><font color='{status_color}'>{rec.get_status_display()}</font></b>"
            
            row = [
                Paragraph(rec.student.student_id, table_cell_center_style),
                Paragraph(rec.student.full_name, table_cell_style),
                Paragraph(rec.student.section.grade.name, table_cell_center_style),
                Paragraph(rec.student.section.name, table_cell_center_style),
                Paragraph(rec.date.strftime('%Y-%m-%d'), table_cell_center_style),
                Paragraph(status_text, table_cell_center_style),
                Paragraph(timestamp_str, table_cell_center_style)
            ]
            table_data.append(row)

        # Page size is 612 x 792 points. Effective printable width is 612 - 72 = 540 points.
        col_widths = [65, 145, 55, 55, 65, 65, 90]
        
        attendance_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Generate Table Styles
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]
        
        # Alternate background colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                t_style.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#F2F5F8')))
                
        attendance_table.setStyle(TableStyle(t_style))
        elements.append(attendance_table)

        # Build Document
        doc.build(elements)
        return response
